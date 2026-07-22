"""Generates output/peer_comparison.xlsx — 11 sheets (one per peer group), with percentile-rank
color coding, benchmark company highlighting, and a summary median row.
"""

import sqlite3
import os
import sys
import pandas as pd
from openpyxl.styles import PatternFill, Font

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "analytics"))
from peer import load_latest_ratios, load_peer_groups, compute_peer_percentiles, METRICS

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "nifty100.db")
DATA_RAW = os.path.join(os.path.dirname(__file__), "..", "data", "raw")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "output")

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
BOLD = Font(bold=True)


def percentile_fill(pct: float):
    if pct is None or pd.isna(pct):
        return None
    if pct >= 0.75:
        return GREEN
    elif pct >= 0.25:
        return YELLOW
    else:
        return RED


def build_peer_comparison_excel():
    conn = sqlite3.connect(DB_PATH)
    ratios = load_latest_ratios(conn)
    peer_groups = load_peer_groups()
    percentiles = compute_peer_percentiles(ratios, peer_groups)

    companies = pd.read_excel(os.path.join(DATA_RAW, "companies.xlsx"), header=1)
    companies["id"] = companies["id"].str.strip().str.upper()
    name_map = dict(zip(companies["id"], companies["company_name"]))

    path = os.path.join(OUTPUT_DIR, "peer_comparison.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for group_name, grp in peer_groups.groupby("peer_group_name"):
            members = grp["company_id"].tolist()
            benchmark = grp.loc[grp["is_benchmark"] == True, "company_id"]
            benchmark_id = benchmark.iloc[0] if len(benchmark) else None

            group_ratios = ratios[ratios["company_id"].isin(members)].copy()
            group_ratios["company_name"] = group_ratios["company_id"].map(name_map)

            pct_pivot = percentiles[percentiles["peer_group_name"] == group_name].pivot(
                index="company_id", columns="metric", values="percentile_rank"
            )
            pct_pivot.columns = [f"{c}_percentile" for c in pct_pivot.columns]

            sheet_df = group_ratios.merge(pct_pivot, on="company_id", how="left")

            display_cols = ["company_id", "company_name"] + METRICS + [f"{m}_percentile" for m in METRICS]
            display_cols = [c for c in display_cols if c in sheet_df.columns]
            sheet_df = sheet_df[display_cols]

            median_row = {"company_id": "SECTOR MEDIAN", "company_name": ""}
            for m in METRICS:
                if m in sheet_df.columns:
                    median_row[m] = sheet_df[m].median()
                pct_col = f"{m}_percentile"
                if pct_col in sheet_df.columns:
                    median_row[pct_col] = None
            sheet_df = pd.concat([sheet_df, pd.DataFrame([median_row])], ignore_index=True)

            safe_name = group_name[:31]
            sheet_df.to_excel(writer, sheet_name=safe_name, index=False)

    _apply_formatting(path, peer_groups, METRICS)
    return path


def _apply_formatting(path, peer_groups, metrics):
    from openpyxl import load_workbook
    wb = load_workbook(path)

    for group_name, grp in peer_groups.groupby("peer_group_name"):
        sheet_name = group_name[:31]
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        benchmark = grp.loc[grp["is_benchmark"] == True, "company_id"]
        benchmark_id = benchmark.iloc[0] if len(benchmark) else None

        cid_col = headers.index("company_id") + 1 if "company_id" in headers else 1

        for row in range(2, ws.max_row + 1):
            cid_cell = ws.cell(row=row, column=cid_col)
            if cid_cell.value == benchmark_id:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = GOLD
            if cid_cell.value == "SECTOR MEDIAN":
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).font = BOLD

            for metric in metrics:
                pct_col_name = f"{metric}_percentile"
                if pct_col_name not in headers:
                    continue
                col_idx = headers.index(pct_col_name) + 1
                cell = ws.cell(row=row, column=col_idx)
                fill = percentile_fill(cell.value)
                if fill:
                    cell.fill = fill

    wb.save(path)


if __name__ == "__main__":
    path = build_peer_comparison_excel()
    print(f"Exported: {path}")

    import openpyxl
    wb = openpyxl.load_workbook(path)
    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
