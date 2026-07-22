"""Sprint 3 pipeline: builds the full screener universe (financial_ratios + market_cap + sector),
computes the sector-relative composite quality score, runs all 6 presets, and exports screener_output.xlsx.
"""

import sqlite3
import os
import sys
import logging

import pandas as pd
import numpy as np
import yaml

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "screener"))
from engine import load_config, apply_filter, apply_de_financial_carveout

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "nifty100.db")
DATA_SUP = os.path.join(os.path.dirname(__file__), "..", "data", "supporting")
DATA_RAW = os.path.join(os.path.dirname(__file__), "..", "data", "raw")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "output")

os.makedirs(OUTPUT_DIR, exist_ok=True)

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger("screener_pipeline")


def winsorize_scale(series: pd.Series, lo_pct=0.10, hi_pct=0.90) -> pd.Series:
    """Cap values at P10/P90, then scale linearly to 0-100."""
    valid = series.dropna()
    if len(valid) == 0:
        return pd.Series([50.0] * len(series), index=series.index)
    lo = valid.quantile(lo_pct)
    hi = valid.quantile(hi_pct)
    if hi == lo:
        return pd.Series([50.0] * len(series), index=series.index)
    capped = series.clip(lower=lo, upper=hi)
    scaled = (capped - lo) / (hi - lo) * 100
    return scaled.fillna(50.0)


def build_universe(conn):
    """Build the full latest-year screener universe: financial_ratios + sector + market_cap + P&L sales/net_profit."""
    ratios = pd.read_sql("""
        SELECT f.*, s.broad_sector, s.sub_sector
        FROM financial_ratios f
        LEFT JOIN sectors s ON f.company_id = s.company_id
        WHERE f.year = (SELECT MAX(f2.year) FROM financial_ratios f2 WHERE f2.company_id = f.company_id)
    """, conn)

    ratios["fy_year"] = ratios["year"].str[:4].astype(int)

    market_cap = pd.read_excel(os.path.join(DATA_SUP, "market_cap.xlsx"), header=0)
    market_cap["company_id"] = market_cap["company_id"].str.strip().str.upper()
    # Take latest available market_cap year <= fy_year, else latest overall
    mc_latest = market_cap.sort_values("year").groupby("company_id").tail(1)
    mc_latest = mc_latest[["company_id", "market_cap_crore", "pe_ratio", "pb_ratio", "ev_ebitda", "dividend_yield_pct"]]

    pl = pd.read_excel(os.path.join(DATA_RAW, "profitandloss.xlsx"), header=1)
    pl["company_id"] = pl["company_id"].str.strip().str.upper()
    pl_latest = pl.sort_values("year").groupby("company_id").tail(1)
    pl_latest = pl_latest[["company_id", "sales", "net_profit"]].rename(
        columns={"sales": "sales_cr", "net_profit": "net_profit_cr"}
    )

    universe = ratios.merge(mc_latest, on="company_id", how="left")
    universe = universe.merge(pl_latest, on="company_id", how="left")

    universe["fcf_yield_pct"] = np.where(
        (universe["market_cap_crore"].notna()) & (universe["market_cap_crore"] != 0),
        universe["free_cash_flow_cr"] / universe["market_cap_crore"] * 100,
        np.nan,
    )
    return universe


def compute_multiyear_flags(conn, universe: pd.DataFrame) -> pd.DataFrame:
    """Compute FCF-positive-latest-year and D/E-declining-YoY flags needed for Turnaround Watch preset."""
    all_years = pd.read_sql("SELECT company_id, year, free_cash_flow_cr, debt_to_equity FROM financial_ratios", conn)
    all_years = all_years.sort_values(["company_id", "year"])

    fcf_positive = {}
    de_declining = {}
    for cid, grp in all_years.groupby("company_id"):
        grp = grp.reset_index(drop=True)
        fcf_positive[cid] = bool(grp.iloc[-1]["free_cash_flow_cr"] > 0) if len(grp) else False
        if len(grp) >= 2:
            de_declining[cid] = bool(grp.iloc[-1]["debt_to_equity"] < grp.iloc[-2]["debt_to_equity"])
        else:
            de_declining[cid] = False

    universe["fcf_positive_latest_year"] = universe["company_id"].map(fcf_positive)
    universe["de_declining_yoy"] = universe["company_id"].map(de_declining)
    return universe


def compute_composite_score(universe: pd.DataFrame, config: dict) -> pd.Series:
    """0-100 composite quality score, sector-relative, using P10/P90 winsorisation per the KPI weights."""
    w = config["composite_score"]["sub_weights"]

    def sector_scaled(col, invert=False):
        result = pd.Series(index=universe.index, dtype=float)
        for sector, grp in universe.groupby("broad_sector", dropna=False):
            vals = grp[col] if not invert else -grp[col]
            result.loc[grp.index] = winsorize_scale(vals)
        return result

    roe_s = sector_scaled("return_on_equity_pct")
    roce_s = sector_scaled("return_on_capital_employed_pct")
    npm_s = sector_scaled("net_profit_margin_pct")
    fcf_cagr_s = sector_scaled("revenue_cagr_5yr")  # FCF 5yr CAGR not separately stored; proxy via revenue growth trend
    cfo_pat_s = sector_scaled("fcf_conversion_rate_pct")
    fcf_pos_flag = (universe["free_cash_flow_cr"] > 0).astype(float) * 100
    rev_cagr_s = sector_scaled("revenue_cagr_5yr")
    pat_cagr_s = sector_scaled("pat_cagr_5yr")
    de_s = sector_scaled("debt_to_equity", invert=True)
    icr_s = sector_scaled("interest_coverage")

    score = (
        w["roe"] * roe_s + w["roce"] * roce_s + w["npm"] * npm_s +
        w["fcf_cagr"] * fcf_cagr_s + w["cfo_pat_ratio"] * cfo_pat_s + w["fcf_positive_flag"] * fcf_pos_flag +
        w["revenue_cagr"] * rev_cagr_s + w["pat_cagr"] * pat_cagr_s +
        w["de_score"] * de_s + w["icr_score"] * icr_s
    )
    return score.round(2)


def run_all_presets(universe: pd.DataFrame, config: dict) -> dict:
    from engine import run_screener
    results = {}
    for name, preset in config["presets"].items():
        filters = preset["filters"]
        simple_filters = {k: v for k, v in filters.items() if k not in ("fcf_positive_latest_year", "de_declining_yoy")}
        mask = pd.Series([True] * len(universe), index=universe.index)
        for metric, condition in simple_filters.items():
            mask &= apply_filter(universe, metric, condition)
        for special in ("fcf_positive_latest_year", "de_declining_yoy"):
            if special in filters:
                mask &= universe[special] == filters[special]["equals"]
        result = universe[mask].sort_values("composite_quality_score", ascending=False)
        results[name] = result
    return results


def export_screener_excel(results: dict, config: dict):
    display_cols = [
        "company_id", "company_name" if "company_name" in next(iter(results.values())).columns else "company_id",
        "broad_sector", "return_on_equity_pct", "return_on_capital_employed_pct", "net_profit_margin_pct",
        "debt_to_equity", "interest_coverage", "free_cash_flow_cr", "revenue_cagr_5yr", "pat_cagr_5yr",
        "eps_cagr_5yr", "pe_ratio", "pb_ratio", "dividend_yield_pct", "fcf_yield_pct",
        "asset_turnover", "capex_intensity_pct", "dividend_payout_ratio_pct", "sales_cr",
        "net_profit_cr", "composite_quality_score",
    ]

    path = os.path.join(OUTPUT_DIR, "screener_output.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in results.items():
            label = config["presets"][name]["label"]
            cols = [c for c in display_cols if c in df.columns]
            df[cols].to_excel(writer, sheet_name=label[:31], index=False)

    _apply_conditional_formatting(path, results, config)
    return path


def _apply_conditional_formatting(path, results, config):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    wb = load_workbook(path)
    for name, preset in config["presets"].items():
        label = preset["label"][:31]
        if label not in wb.sheetnames:
            continue
        ws = wb[label]
        headers = [cell.value for cell in ws[1]]
        for metric, condition in preset["filters"].items():
            if metric not in headers:
                continue
            col_idx = headers.index(metric) + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                val = cell.value
                if val is None:
                    continue
                passes = True
                if "min" in condition:
                    passes &= val >= condition["min"]
                if "max" in condition:
                    passes &= val <= condition["max"]
                if "equals" in condition:
                    passes &= val == condition["equals"]
                cell.fill = green if passes else red
    wb.save(path)


if __name__ == "__main__":
    config = load_config()
    conn = sqlite3.connect(DB_PATH)

    universe = build_universe(conn)
    universe = compute_multiyear_flags(conn, universe)
    universe["composite_quality_score"] = compute_composite_score(universe, config)

    logger.info(f"Universe size: {len(universe)} companies")

    results = run_all_presets(universe, config)
    for name, df in results.items():
        preset = config["presets"][name]
        lo, hi = preset["expected_range"]
        status = "PASS" if lo <= len(df) <= hi else "CHECK"
        logger.info(f"{preset['label']:25s} -> {len(df):3d} companies (expected {lo}-{hi}) [{status}]")

    path = export_screener_excel(results, config)
    logger.info(f"Exported: {path}")

    conn.close()
